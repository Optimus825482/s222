"""
Export module — MD, Excel, PDF export for task results.
"""

from __future__ import annotations

import io
import os
import time
from typing import Any

import streamlit as st

from core.models import Thread, Task


def render_export_buttons(result: str, task: Task | None = None) -> None:
    """Render export buttons below a result."""
    if not result or len(result.strip()) < 10:
        return

    cols = st.columns([1, 1, 1, 4])

    with cols[0]:
        md_bytes = _export_markdown(result, task)
        st.download_button(
            "📄 MD",
            data=md_bytes,
            file_name=f"result_{int(time.time())}.md",
            mime="text/markdown",
            key=f"export_md_{id(result)}_{time.time_ns()}",
            use_container_width=True,
        )

    with cols[1]:
        xlsx_bytes = _export_excel(result, task)
        st.download_button(
            "📊 Excel",
            data=xlsx_bytes,
            file_name=f"result_{int(time.time())}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"export_xlsx_{id(result)}_{time.time_ns()}",
            use_container_width=True,
        )

    with cols[2]:
        pdf_bytes = _export_pdf(result, task)
        st.download_button(
            "📕 PDF",
            data=pdf_bytes,
            file_name=f"result_{int(time.time())}.pdf",
            mime="application/pdf",
            key=f"export_pdf_{id(result)}_{time.time_ns()}",
            use_container_width=True,
        )


def _export_markdown(result: str, task: Task | None) -> bytes:
    """Generate markdown export."""
    lines = ["# Multi-Agent Result\n"]
    if task:
        lines.append(f"**Query:** {task.user_input}\n")
        lines.append(f"**Pipeline:** {task.pipeline_type.value}\n")
        if task.sub_tasks:
            lines.append(f"**Agents:** {', '.join(sub.assigned_agent.value for sub in task.sub_tasks)}\n")
        if task.total_latency_ms:
            lines.append(f"**Latency:** {task.total_latency_ms:.0f}ms\n")
        lines.append("---\n")
    lines.append(result)
    return "\n".join(lines).encode("utf-8")


def _export_excel(result: str, task: Task | None) -> bytes:
    """Generate Excel export with openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    row = 1
    if task:
        for label, value in [
            ("Query", task.user_input),
            ("Pipeline", task.pipeline_type.value),
            ("Agents", ", ".join(sub.assigned_agent.value for sub in task.sub_tasks) if task.sub_tasks else "direct"),
            ("Latency", f"{task.total_latency_ms:.0f}ms" if task.total_latency_ms else "—"),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Result").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 1

    for line in result.split("\n"):
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_pdf(result: str, task: Task | None) -> bytes:
    """Generate professional research report PDF with Turkish character support."""
    from datetime import datetime, timezone
    from pathlib import Path as _Path
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    import reportlab as _rl

    # ── Register a Unicode-capable font for Turkish chars ──
    _FONT = "Helvetica"
    _FONT_BOLD = "Helvetica-Bold"
    _rl_fonts_dir = _Path(_rl.__file__).parent / "fonts"
    _font_candidates = [
        # reportlab bundled Vera — always available, full Unicode/Turkish
        ("VeraUI", str(_rl_fonts_dir / "Vera.ttf"), str(_rl_fonts_dir / "VeraBd.ttf")),
        # Windows
        ("WinArial", "C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("WinCalibri", "C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
        ("WinSegoe", "C:/Windows/Fonts/segoeui.ttf", "C:/Windows/Fonts/segoeuib.ttf"),
        # Linux / Docker
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("Liberation", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
         "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
    ]
    for fname, regular, bold in _font_candidates:
        if not os.path.exists(regular):
            continue
        try:
            pdfmetrics.registerFont(TTFont(fname, regular))
            bold_name = f"{fname}-Bold"
            if os.path.exists(bold):
                pdfmetrics.registerFont(TTFont(bold_name, bold))
            else:
                pdfmetrics.registerFont(TTFont(bold_name, regular))
            _FONT = fname
            _FONT_BOLD = bold_name
            break
        except Exception:
            continue

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=2.5 * cm,
        bottomMargin=2 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
    )

    COLOR_PRIMARY = HexColor("#1e3a5f")
    COLOR_ACCENT = HexColor("#3b82f6")
    COLOR_TEXT = HexColor("#1f2937")
    COLOR_MUTED = HexColor("#6b7280")
    COLOR_LIGHT_BG = HexColor("#f0f4f8")

    s_title = ParagraphStyle(
        "ReportTitle", fontName=_FONT_BOLD, fontSize=22, leading=28,
        textColor=COLOR_PRIMARY, alignment=TA_CENTER, spaceAfter=6,
    )
    s_subtitle = ParagraphStyle(
        "ReportSubtitle", fontName=_FONT, fontSize=11, leading=14,
        textColor=COLOR_MUTED, alignment=TA_CENTER, spaceAfter=20,
    )
    s_section = ParagraphStyle(
        "SectionHead", fontName=_FONT_BOLD, fontSize=13, leading=18,
        textColor=COLOR_PRIMARY, spaceBefore=16, spaceAfter=8,
    )
    s_meta_label = ParagraphStyle(
        "MetaLabel", fontName=_FONT_BOLD, fontSize=9, leading=12,
        textColor=COLOR_MUTED,
    )
    s_meta_value = ParagraphStyle(
        "MetaValue", fontName=_FONT, fontSize=10, leading=14,
        textColor=COLOR_TEXT,
    )
    s_body = ParagraphStyle(
        "Body", fontName=_FONT, fontSize=10.5, leading=16,
        textColor=COLOR_TEXT, alignment=TA_JUSTIFY, spaceAfter=8,
    )
    s_footer = ParagraphStyle(
        "Footer", fontName=_FONT, fontSize=8, leading=10,
        textColor=COLOR_MUTED, alignment=TA_CENTER,
    )

    story = []
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph("Multi-Agent Araştırma Raporu", s_title))

    now_str = datetime.now(timezone.utc).strftime("%d %B %Y, %H:%M UTC")
    story.append(Paragraph(f"Oluşturulma: {now_str}", s_subtitle))
    story.append(HRFlowable(
        width="100%", thickness=2, color=COLOR_ACCENT,
        spaceAfter=16, spaceBefore=4,
    ))

    if task:
        agents_str = ", ".join(
            s.assigned_agent.value for s in task.sub_tasks
        ) if task.sub_tasks else "direct"
        latency_str = f"{task.total_latency_ms:.0f}ms" if task.total_latency_ms else "—"

        meta_data = [
            [Paragraph("SORGU", s_meta_label), Paragraph(_pdf_escape(task.user_input), s_meta_value)],
            [Paragraph("PIPELINE", s_meta_label), Paragraph(task.pipeline_type.value.upper(), s_meta_value)],
            [Paragraph("AGENT'LAR", s_meta_label), Paragraph(agents_str, s_meta_value)],
            [Paragraph("SÜRE", s_meta_label), Paragraph(latency_str, s_meta_value)],
        ]
        meta_table = Table(meta_data, colWidths=[3 * cm, 12.5 * cm])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), COLOR_LIGHT_BG),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("LINEBELOW", (0, 0), (-1, -2), 0.5, HexColor("#e5e7eb")),
            ("BOX", (0, 0), (-1, -1), 0.5, HexColor("#d1d5db")),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 16))

    story.append(HRFlowable(
        width="100%", thickness=0.5, color=HexColor("#e5e7eb"),
        spaceAfter=12, spaceBefore=4,
    ))
    story.append(Paragraph("Analiz Sonuçları", s_section))

    for para in result.split("\n\n"):
        text = para.strip()
        if not text:
            continue
        if text.startswith(("### ", "## ", "# ")):
            clean = text.lstrip("#").strip().replace("**", "")
            story.append(Paragraph(_pdf_escape(clean), s_section))
        elif text.startswith("---"):
            story.append(HRFlowable(
                width="100%", thickness=0.5, color=HexColor("#d1d5db"),
                spaceAfter=8, spaceBefore=8,
            ))
        else:
            processed = _md_bold_to_pdf(_pdf_escape(text))
            story.append(Paragraph(processed, s_body))

    story.append(Spacer(1, 2 * cm))
    story.append(HRFlowable(
        width="100%", thickness=0.5, color=HexColor("#e5e7eb"),
        spaceAfter=8,
    ))
    story.append(Paragraph(
        "Bu rapor Multi-Agent Operations Center tarafından otomatik oluşturulmuştur.",
        s_footer,
    ))

    doc.build(story)
    return buf.getvalue()


def _pdf_escape(text: str) -> str:
    """Escape special chars for reportlab XML."""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


def _md_bold_to_pdf(text: str) -> str:
    """Convert markdown **bold** to reportlab <b>bold</b>."""
    import re
    return re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
